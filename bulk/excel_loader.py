"""Small Excel loader for bulk posting rows.

This replaces the old pandas-only read path for the simple worksheet shape
used by AIMAX bulk posting.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


COL_ID = ["아이디", "id", "계정", "account"]
COL_PW = ["비밀번호", "password", "pw", "패스워드"]
COL_KW = ["핵심키워드", "키워드", "keyword", "kw", "주제"]


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def find_col(columns: list[str], candidates: list[str]) -> str | None:
    lower = {str(column).strip().lower(): str(column).strip() for column in columns if str(column).strip()}
    for candidate in candidates:
        matched = lower.get(candidate.lower())
        if matched:
            return matched
    return None


def load_bulk_rows(excel_path: str | Path) -> list[dict[str, str]]:
    path = Path(excel_path)
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("대량 발행 엑셀은 .xlsx 또는 .xlsm 형식으로 저장해주세요.")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        try:
            header_values = next(iterator)
        except StopIteration:
            raise ValueError("엑셀 파일에 헤더 행이 없습니다.") from None

        columns = [normalize_cell(value) for value in header_values]
        col_id = find_col(columns, COL_ID)
        col_pw = find_col(columns, COL_PW)
        col_kw = find_col(columns, COL_KW)
        if not col_id or not col_pw or not col_kw:
            missing = []
            if not col_id:
                missing.append("아이디")
            if not col_pw:
                missing.append("비밀번호")
            if not col_kw:
                missing.append("핵심키워드")
            raise ValueError(
                f"엑셀 열을 찾을 수 없습니다: {', '.join(missing)}\n"
                f"현재 열 이름: {columns}"
            )

        index_by_column = {column: index for index, column in enumerate(columns)}
        rows: list[dict[str, str]] = []
        for row_values in iterator:
            values = [normalize_cell(value) for value in row_values]
            if not any(values):
                continue
            row = {
                "account_id": values[index_by_column[col_id]] if index_by_column[col_id] < len(values) else "",
                "account_pw": values[index_by_column[col_pw]] if index_by_column[col_pw] < len(values) else "",
                "keyword": values[index_by_column[col_kw]] if index_by_column[col_kw] < len(values) else "",
            }
            if not any(row.values()):
                continue
            rows.append(row)
        return rows
    finally:
        workbook.close()
